import openpyxl, os, pprint
os.chdir("C:\\Users\\sam\\Downloads")
wb = openpyxl.load_workbook('student.xlsx')
ws = wb['studentActivityPointsReport']
stds = {}
for i in range(3, 55):
	stds[ws['A' + str(i)].value] = ws['B'  + str(i)].value
pprint.pprint(stds)
ws['A1'].value = "Activity Points"
wb.save("new.xlsx")
wb.create_sheet()
wb['Sheet'].title = "AdmissionNumber"
ws = wb['AdmissionNumber']
i = 1
for key in stds:
	ws["A" + str(i)] = key
	ws["B" + str(i)] = stds[key]
	i = i + 1
wb.save("new.xlsx")